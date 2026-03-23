# -*- coding: utf-8 -*-

# === CONFIGURACIÓN DE RUTAS ===
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# ===============================

"""
Módulo para exportar datos de proveedores a Excel con informes y gráficos
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
import sqlite3
from typing import List, Dict
import os


class ExportadorExcel:
    """Clase para exportar datos a Excel con formato profesional."""
    
    def __init__(self, db_path='facturas.db'):
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path)
        self.conn.row_factory = sqlite3.Row
        self.cursor = self.conn.cursor()
    
    def exportar_proveedores_completo(self, ruta_salida=None):
        """
        Exporta un archivo Excel completo con:
        - Lista de proveedores
        - Facturas por proveedor
        - Análisis de gastos
        - Gráficos
        """
        if not ruta_salida:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta_salida = f"Informe_Proveedores_{timestamp}.xlsx"
        
        wb = openpyxl.Workbook()
        
        # Eliminar hoja por defecto
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Crear hojas
        self._crear_hoja_proveedores(wb)
        self._crear_hoja_facturas(wb)
        self._crear_hoja_analisis_gastos(wb)
        self._crear_hoja_comparativa_mensual(wb)
        
        # Guardar
        wb.save(ruta_salida)
        self.conn.close()
        
        return ruta_salida
    
    def _crear_hoja_proveedores(self, wb):
        """Crea hoja con lista de proveedores."""
        ws = wb.create_sheet("Proveedores", 0)
        
        # Títulos
        headers = ["ID", "Nombre", "Nº Proveedor", "CIF/NIF", "Cuenta Gasto", 
                   "Categoría", "Email", "IBAN", "Fecha Creación"]
        
        # Estilo del encabezado
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        self.cursor.execute("""
            SELECT id, nombre, numero_proveedor, cif_nif, cuenta_gasto, 
                   categoria, email, iban, fecha_creacion 
            FROM proveedores 
            ORDER BY nombre
        """)
        
        proveedores = self.cursor.fetchall()
        
        for row_idx, prov in enumerate(proveedores, 2):
            for col_idx, value in enumerate(prov, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Alternar colores de fila
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Ajustar anchos de columna
        column_widths = [8, 30, 15, 15, 15, 25, 30, 25, 20]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Congelar primera fila
        ws.freeze_panes = "A2"
        
        # Agregar filtros
        ws.auto_filter.ref = ws.dimensions
    
    def _crear_hoja_facturas(self, wb):
        """Crea hoja con todas las facturas procesadas."""
        ws = wb.create_sheet("Facturas Procesadas")
        
        # Títulos
        headers = ["Fecha", "Proveedor", "Nº Factura", "Tipo", "Serie", 
                   "Cuenta Gasto", "Archivo"]
        
        # Estilo del encabezado
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        self.cursor.execute("""
            SELECT fecha_procesado, nombre_proveedor, numero_factura, 
                   tipo_factura, serie_factura, cuenta_gasto, nombre_archivo
            FROM historial_procesado 
            ORDER BY fecha_procesado DESC
        """)
        
        facturas = self.cursor.fetchall()
        
        for row_idx, factura in enumerate(facturas, 2):
            for col_idx, value in enumerate(factura, 1):
                # Formatear fecha si es la primera columna
                if col_idx == 1 and value:
                    try:
                        fecha = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
                        value = fecha.strftime("%d/%m/%Y")
                    except Exception as _e:
                        pass  # valor no es fecha, se usa tal cual
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Alternar colores de fila
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Ajustar anchos
        column_widths = [12, 25, 18, 12, 10, 15, 35]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Congelar primera fila
        ws.freeze_panes = "A2"
        
        # Agregar filtros
        ws.auto_filter.ref = ws.dimensions
    
    def _crear_hoja_analisis_gastos(self, wb):
        """Crea hoja con análisis de gastos por proveedor."""
        ws = wb.create_sheet("Análisis por Proveedor")
        
        # Título
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "ANÁLISIS DE FACTURAS POR PROVEEDOR"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = ["Proveedor", "Total Facturas", "Última Factura", "Cuenta Principal", "Categoría"]
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Obtener datos agrupados
        self.cursor.execute("""
            SELECT 
                h.nombre_proveedor,
                COUNT(*) as total_facturas,
                MAX(h.fecha_procesado) as ultima_factura,
                h.cuenta_gasto,
                p.categoria
            FROM historial_procesado h
            LEFT JOIN proveedores p ON h.proveedor_id = p.id
            GROUP BY h.nombre_proveedor
            ORDER BY total_facturas DESC
        """)
        
        datos = self.cursor.fetchall()
        
        for row_idx, dato in enumerate(datos, 3):
            proveedor, total, ultima, cuenta, categoria = dato
            
            # Formatear fecha
            if ultima:
                try:
                    fecha = datetime.strptime(ultima, "%Y-%m-%d %H:%M:%S")
                    ultima = fecha.strftime("%d/%m/%Y")
                except Exception as _e:
                    pass  # valor no es fecha, se usa tal cual
            
            ws.cell(row=row_idx, column=1, value=proveedor)
            ws.cell(row=row_idx, column=2, value=total)
            ws.cell(row=row_idx, column=3, value=ultima)
            ws.cell(row=row_idx, column=4, value=cuenta)
            ws.cell(row=row_idx, column=5, value=categoria)
            
            # Alternar colores
            if row_idx % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 25
        
        # Crear gráfico de barras
        if len(datos) > 0:
            chart = BarChart()
            chart.title = "Facturas por Proveedor"
            chart.style = 10
            chart.y_axis.title = "Número de Facturas"
            chart.x_axis.title = "Proveedor"
            
            data = Reference(ws, min_col=2, min_row=2, max_row=len(datos)+2)
            cats = Reference(ws, min_col=1, min_row=3, max_row=len(datos)+2)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "G3")
    
    def _crear_hoja_comparativa_mensual(self, wb):
        """Crea hoja con comparativa de gastos por mes."""
        ws = wb.create_sheet("Comparativa Mensual")
        
        # Título
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "COMPARATIVA DE FACTURAS POR MES"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = ["Año", "Mes", "Total Facturas", "Proveedores Únicos"]
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Obtener datos por mes
        self.cursor.execute("""
            SELECT 
                strftime('%Y', fecha_procesado) as año,
                strftime('%m', fecha_procesado) as mes,
                COUNT(*) as total_facturas,
                COUNT(DISTINCT nombre_proveedor) as proveedores_unicos
            FROM historial_procesado
            WHERE fecha_procesado IS NOT NULL
            GROUP BY año, mes
            ORDER BY año DESC, mes DESC
        """)
        
        datos = self.cursor.fetchall()
        
        meses = {
            '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
            '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
            '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
        }
        
        for row_idx, dato in enumerate(datos, 3):
            año, mes, total, proveedores = dato
            nombre_mes = meses.get(mes, mes)
            
            ws.cell(row=row_idx, column=1, value=año)
            ws.cell(row=row_idx, column=2, value=nombre_mes)
            ws.cell(row=row_idx, column=3, value=total)
            ws.cell(row=row_idx, column=4, value=proveedores)
            
            # Alternar colores
            if row_idx % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                    )
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 20
        
        # Crear gráfico de líneas
        if len(datos) > 0:
            chart = LineChart()
            chart.title = "Evolución de Facturas por Mes"
            chart.style = 10
            chart.y_axis.title = "Cantidad"
            chart.x_axis.title = "Mes"
            
            data = Reference(ws, min_col=3, min_row=2, max_row=len(datos)+2, max_col=4)
            cats = Reference(ws, min_col=2, min_row=3, max_row=len(datos)+2)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "F3")
    
    def exportar_proveedor_individual(self, proveedor_id, ruta_salida=None):
        """Exporta datos de un proveedor específico con sus facturas."""
        # Obtener datos del proveedor
        self.cursor.execute("SELECT * FROM proveedores WHERE id = ?", (proveedor_id,))
        proveedor = self.cursor.fetchone()
        
        if not proveedor:
            return None
        
        if not ruta_salida:
            nombre_prov = proveedor['nombre'].replace(' ', '_')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta_salida = f"Proveedor_{nombre_prov}_{timestamp}.xlsx"
        
        wb = openpyxl.Workbook()
        
        # Eliminar hoja por defecto
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Hoja de datos del proveedor
        ws_datos = wb.create_sheet("Datos del Proveedor", 0)
        
        # Título
        ws_datos['A1'] = "DATOS DEL PROVEEDOR"
        ws_datos['A1'].font = Font(bold=True, size=14)
        
        # Datos
        datos_prov = [
            ("Nombre:", proveedor['nombre']),
            ("Nº Proveedor:", proveedor['numero_proveedor']),
            ("CIF/NIF:", proveedor['cif_nif']),
            ("Cuenta Gasto:", proveedor['cuenta_gasto']),
            ("Categoría:", proveedor['categoria']),
            ("Email:", proveedor['email']),
            ("IBAN:", proveedor['iban']),
            ("Fecha Creación:", proveedor['fecha_creacion'])
        ]
        
        for row_idx, (label, value) in enumerate(datos_prov, 3):
            ws_datos.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_datos.cell(row=row_idx, column=2, value=value)
        
        ws_datos.column_dimensions['A'].width = 20
        ws_datos.column_dimensions['B'].width = 40
        
        # Hoja de facturas del proveedor
        ws_facturas = wb.create_sheet("Facturas")
        
        headers = ["Fecha", "Nº Factura", "Tipo", "Serie", "Cuenta Gasto", "Archivo"]
        for col, header in enumerate(headers, 1):
            cell = ws_facturas.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Obtener facturas del proveedor
        self.cursor.execute("""
            SELECT fecha_procesado, numero_factura, tipo_factura, 
                   serie_factura, cuenta_gasto, nombre_archivo
            FROM historial_procesado
            WHERE proveedor_id = ?
            ORDER BY fecha_procesado DESC
        """, (proveedor_id,))
        
        facturas = self.cursor.fetchall()
        
        for row_idx, factura in enumerate(facturas, 2):
            for col_idx, value in enumerate(factura, 1):
                ws_facturas.cell(row=row_idx, column=col_idx, value=value)
        
        column_widths = [18, 18, 12, 10, 15, 35]
        for col_idx, width in enumerate(column_widths, 1):
            ws_facturas.column_dimensions[get_column_letter(col_idx)].width = width
        
        wb.save(ruta_salida)
        self.conn.close()
        
        return ruta_salida


# Función helper para usar desde la interfaz
def exportar_todos_proveedores(ruta_salida=None):
    """Función helper para exportar todos los proveedores."""
    exportador = ExportadorExcel()
    return exportador.exportar_proveedores_completo(ruta_salida)


def exportar_un_proveedor(proveedor_id, ruta_salida=None):
    """Función helper para exportar un proveedor específico."""
    exportador = ExportadorExcel()
    return exportador.exportar_proveedor_individual(proveedor_id, ruta_salida)
