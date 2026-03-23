# -*- coding: utf-8 -*-
"""
Excel Contable — Asientos Contables Profesionales — V11.
Genera asientos contables en formato libro diario:
  - Cuenta de gasto (DEBE): cuenta contable del gasto (con subcuenta)
  - IVA soportado (DEBE): cuenta 472000
  - Proveedor (HABER): cuenta de proveedor (con subcuenta)
CORREGIDO: Una sola columna "Cuenta" y una sola columna "Subcuenta"
"""
from __future__ import annotations
from datetime import datetime
from typing import List, Dict, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


_HDR_FILL  = "1F4E79"
_HDR_FONT  = "FFFFFF"
_DEBE_COLOR = "E6F4EA"   # verde claro para DEBE
_HABER_COLOR = "FFF3E0"  # naranja claro para HABER


def _thin_border():
    s = Side(border_style="thin", color="CBD5E0")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=True, color=_HDR_FONT, size=10)
    cell.fill      = PatternFill("solid", fgColor=_HDR_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()
    return cell


def exportar_contable(
    facturas: List[Dict],
    ruta_salida: Optional[str] = None,
    empresa_nombre: str = "Mi Empresa",
    filtro_cont: str = "todos",   # "todos" | "automatica" | "manual"
    agrupar_por: str = "ninguno", # "ninguno" | "proveedor" | "categoria" | "tipo"
) -> str:
    """
    Genera el libro de asientos contables en Excel.
    filtro_cont filtra por tipo de contabilización.
    agrupar_por genera hojas adicionales agrupadas.
    Cada factura genera hasta 3 líneas de asiento:
      1. DEBE  → cuenta_gasto / subcuenta_gasto por la base imponible
      2. DEBE  → 472000 (IVA soportado) por el IVA
      3. HABER → cuenta_proveedor / subcuenta_proveedor por el total
    """
    if not OPENPYXL_OK:
        raise ImportError("openpyxl no instalado. pip install openpyxl")

    if not ruta_salida:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_salida = f"Contabilidad_{ts}.xlsx"

    # Aplicar filtro de contabilización
    if filtro_cont == "automatica":
        facturas = [f for f in facturas if f.get("cont_automatica", 0)]
    elif filtro_cont == "manual":
        facturas = [f for f in facturas if not f.get("cont_automatica", 0)]

    wb = openpyxl.Workbook()

    # ── Hoja Libro Diario ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Libro Diario"

    # Título
    ws.merge_cells("A1:K1")
    tc = ws["A1"]
    tc.value     = f"📘  LIBRO DIARIO — {empresa_nombre}"
    tc.font      = Font(bold=True, size=14, color=_HDR_FONT)
    tc.fill      = PatternFill("solid", fgColor="2B6CB0")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    sc = ws["A2"]
    sc.value     = f"Ejercicio: {datetime.now().year}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sc.font      = Font(italic=True, size=9, color="4A5568")
    sc.alignment = Alignment(horizontal="right")

    # Cabeceras: una sola columna para cuenta y una para subcuenta
    headers = [
        "Nº Asiento", "Fecha", 
        "Cuenta", "Subcuenta",
        "Descripción", "Concepto / Proveedor", 
        "DEBE (€)", "HABER (€)",
        "CIF Proveedor", "Referencia", "Cont."
    ]
    for col, hdr in enumerate(headers, 1):
        _header_cell(ws, 3, col, hdr)
    ws.row_dimensions[3].height = 24

    asiento_num = 1
    row_idx     = 4
    total_debe  = total_haber = 0.0

    for f in facturas:
        fecha           = f.get("fecha_contabilizacion") or f.get("fecha", "")
        cont_lbl        = "AUTO" if f.get("cont_automatica", 0) else "MANUAL"
        proveedor       = f.get("nombre_proveedor", "DESCONOCIDO")
        cif_prov        = f.get("cif_proveedor") or f.get("cif_prov", "")
        
        # --- USAR LOS CAMPOS REALES DE LA BD ---
        cuenta_gasto    = f.get("cuenta_gasto", "629000") or "629000"
        subcuenta_gasto = f.get("subcuenta_gasto", "") or ""
        
        cuenta_proveedor = f.get("cuenta_proveedor", "400000") or "400000"
        subcuenta_proveedor = f.get("subcuenta_proveedor", "") or ""
        # ---------------------------------------
        
        base            = float(f.get("base_imponible", 0) or 0)
        iva             = float(f.get("iva", 0) or 0)
        total           = float(f.get("total", 0) or 0)
        num_factura     = f.get("numero_factura", "")
        es_rectificativa = bool(f.get("es_rectificativa", 0))
        tipo_iva        = int(f.get("tipo_iva", 21) or 21)
        es_exenta       = (tipo_iva == 0)
        iva_label       = ("IVA Exento" if es_exenta
                           else f"IVA {tipo_iva}% Fra. {num_factura or '—'} / {proveedor}")

        # Si es rectificativa, invertimos signos
        if es_rectificativa:
            base = -base
            iva = -iva
            total = -total

        concepto = f"Fra. {num_factura or '—'} / {proveedor}"

        if not es_rectificativa:
            # ── FACTURA NORMAL (cargo) ─────────────────────────────────────────
            # Línea 1: DEBE cuenta de gasto
            if abs(base) > 0.01:
                _asiento_row(ws, row_idx, asiento_num, fecha,
                             cuenta_gasto, subcuenta_gasto,
                             _desc_cuenta(cuenta_gasto), concepto,
                             debe=base, haber=None, cif=cif_prov, ref=num_factura,
                             color=_DEBE_COLOR)
                total_debe += base
                row_idx += 1
                ws.cell(row=row_idx-1, column=11, value=cont_lbl)

            # Línea 2: DEBE IVA soportado (472000) — omitir si exenta
            if abs(iva) > 0.01 and not es_exenta:
                _asiento_row(ws, row_idx, asiento_num, fecha,
                             "472000", "",
                             "H.P. IVA Soportado",
                             iva_label,
                             debe=iva, haber=None, cif=cif_prov, ref=num_factura,
                             color=_DEBE_COLOR)
                total_debe += iva
                row_idx += 1
                ws.cell(row=row_idx-1, column=11, value=cont_lbl)

            # Línea 3: HABER proveedor/acreedor
            if abs(total) > 0.01:
                _asiento_row(ws, row_idx, asiento_num, fecha,
                             cuenta_proveedor, subcuenta_proveedor,
                             f"Proveedores — {proveedor}",
                             concepto,
                             debe=None, haber=total, cif=cif_prov, ref=num_factura,
                             color=_HABER_COLOR)
                total_haber += total
                row_idx += 1
                ws.cell(row=row_idx-1, column=11, value=cont_lbl)

        else:
            # ── FACTURA DE ABONO / RECTIFICATIVA ──────────────────────────────
            # Inversión: lo que era DEBE pasa a HABER y viceversa.
            # Línea 1: HABER cuenta de gasto (anulación del gasto)
            _base_abs = abs(base)
            _iva_abs  = abs(iva)
            _tot_abs  = abs(total)

            if _base_abs > 0.01:
                _asiento_row(ws, row_idx, asiento_num, fecha,
                             cuenta_gasto, subcuenta_gasto,
                             _desc_cuenta(cuenta_gasto),
                             f"ABONO {concepto}",
                             debe=None, haber=_base_abs, cif=cif_prov, ref=num_factura,
                             color=_HABER_COLOR)
                total_haber += _base_abs
                row_idx += 1
                ws.cell(row=row_idx-1, column=11, value=cont_lbl)

            # Línea 2: HABER IVA soportado (476000 IVA devuelto)
            if _iva_abs > 0.01:
                _asiento_row(ws, row_idx, asiento_num, fecha,
                             "472000", "",
                             "H.P. IVA Soportado (abono)",
                             f"IVA ABONO {num_factura or '—'} / {proveedor}",
                             debe=None, haber=_iva_abs, cif=cif_prov, ref=num_factura,
                             color=_HABER_COLOR)
                total_haber += _iva_abs
                row_idx += 1
                ws.cell(row=row_idx-1, column=11, value=cont_lbl)

            # Línea 3: DEBE proveedor/acreedor (nos debe devolver el dinero)
            if _tot_abs > 0.01:
                _asiento_row(ws, row_idx, asiento_num, fecha,
                             cuenta_proveedor, subcuenta_proveedor,
                             f"Proveedores — {proveedor} (abono)",
                             f"ABONO {concepto}",
                             debe=_tot_abs, haber=None, cif=cif_prov, ref=num_factura,
                             color=_DEBE_COLOR)
                total_debe += _tot_abs
                row_idx += 1
                ws.cell(row=row_idx-1, column=11, value=cont_lbl)

        # Separador visual entre asientos
        asiento_num += 1

    # ── Fila de totales ────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row_idx}:F{row_idx}")
    tot = ws[f"A{row_idx}"]
    tot.value = f"TOTALES  ({asiento_num - 1} asientos)"
    tot.font  = Font(bold=True, color=_HDR_FONT)
    tot.fill  = PatternFill("solid", fgColor=_HDR_FILL)
    tot.alignment = Alignment(horizontal="center")

    for col, val in [(7, total_debe), (8, total_haber)]:
        c = ws.cell(row=row_idx, column=col, value=val)
        c.number_format = '#,##0.00 €'
        c.font = Font(bold=True, color=_HDR_FONT)
        c.fill = PatternFill("solid", fgColor=_HDR_FILL)
        c.alignment = Alignment(horizontal="right")

    # Equilibrio debe/haber
    diff = round(total_debe - total_haber, 2)
    if abs(diff) > 0.01:
        ws.merge_cells(f"A{row_idx+1}:K{row_idx+1}")
        warn = ws[f"A{row_idx+1}"]
        warn.value = f"⚠️ Diferencia debe/haber: {diff:.2f} €  (posibles facturas sin base/IVA/total)"
        warn.font  = Font(bold=True, color="C53030")

    # ── Anchos ────────────────────────────────────────────────────────────────
    widths = [11, 12, 12, 12, 30, 40, 14, 14, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja saldo por cuenta ─────────────────────────────────────────────────
    _add_saldo_cuentas(wb, facturas)

    # Hojas separadas auto/manual si hay mezcla y filtro=todos
    if filtro_cont == "todos":
        tiene_auto   = any(f.get("cont_automatica", 0) for f in facturas)
        tiene_manual = any(not f.get("cont_automatica", 0) for f in facturas)
        if tiene_auto and tiene_manual:
            _build_hoja_cont(wb, [f for f in facturas if     f.get("cont_automatica", 0)],
                             "Cont. Automática", empresa_nombre)
            _build_hoja_cont(wb, [f for f in facturas if not f.get("cont_automatica", 0)],
                             "Cont. Manual", empresa_nombre)

    wb.save(ruta_salida)
    return ruta_salida


def _build_hoja_cont(wb, facturas, titulo, empresa_nombre):
    """Hoja extra: subconjunto de facturas por tipo de contabilización."""
    ws = wb.create_sheet(titulo[:28])
    ws.merge_cells("A1:K1")
    tc = ws["A1"]
    tc.value = f"📘  {titulo.upper()} — {empresa_nombre}"
    tc.font  = Font(bold=True, size=13, color="FFFFFF")
    tc.fill  = PatternFill("solid", fgColor="2B6CB0")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    headers = ["Nº Asiento","Fecha","Cuenta","Subcuenta","Descripción",
               "Concepto / Proveedor","DEBE (€)","HABER (€)","CIF Proveedor","Referencia","Cont."]
    for col, hdr in enumerate(headers, 1):
        _header_cell(ws, 2, col, hdr)
    row_idx = 3
    asiento_num = 1
    for f in facturas:
        cont_lbl = "AUTO" if f.get("cont_automatica", 0) else "MANUAL"
        fecha = f.get("fecha", "")
        prov  = f.get("nombre_proveedor", "DESCONOCIDO")
        cif   = f.get("cif_proveedor") or f.get("cif_prov", "")
        cg    = f.get("cuenta_gasto", "629000") or "629000"
        scg   = f.get("subcuenta_gasto", "") or ""
        cp    = f.get("cuenta_proveedor", "400000") or "400000"
        scp   = f.get("subcuenta_proveedor", "") or ""
        base  = float(f.get("base_imponible", 0) or 0)
        iva   = float(f.get("iva", 0) or 0)
        total = float(f.get("total", 0) or 0)
        nfra  = f.get("numero_factura", "")
        conc  = f"Fra. {nfra or '—'} / {prov}"
        for debe, haber, cuenta, subcta, desc in [
            (base,  None,  cg,       scg, _desc_cuenta(cg)),
            (iva,   None,  "472000", "",  "H.P. IVA Soportado"),
            (None,  total, cp,       scp, f"Proveedores — {prov}"),
        ]:
            importe = debe if debe is not None else haber
            if abs(importe or 0) > 0.01:
                color = _DEBE_COLOR if debe is not None else _HABER_COLOR
                _asiento_row(ws, row_idx, asiento_num, fecha, cuenta, subcta,
                             desc, conc, debe=debe, haber=haber,
                             cif=cif, ref=nfra, color=color)
                ws.cell(row=row_idx, column=11, value=cont_lbl)
                row_idx += 1
        asiento_num += 1
    widths = [11,12,12,12,30,38,14,14,14,16,8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _asiento_row(ws, row, num, fecha,
                 cuenta, subcuenta,
                 desc_cuenta, concepto,
                 debe, haber, cif, ref, color):
    fills = PatternFill("solid", fgColor=color.replace("#", ""))
    bordr = _thin_border()

    def _c(col, val, num_fmt=None, bold=False):
        c = ws.cell(row=row, column=col, value=val)
        c.fill   = fills
        c.border = bordr
        if num_fmt:
            c.number_format = num_fmt
            c.alignment = Alignment(horizontal="right")
        if bold:
            c.font = Font(bold=True)
        return c

    _c(1, num)
    _c(2, fecha)
    _c(3, cuenta or "")
    _c(4, subcuenta or "")
    _c(5, desc_cuenta)
    _c(6, concepto)
    
    if debe is not None:
        _c(7, debe, '#,##0.00 €', bold=True)
        ws.cell(row=row, column=8, value=None).fill = fills
    else:
        ws.cell(row=row, column=7, value=None).fill = fills
        _c(8, haber, '#,##0.00 €', bold=True)
    
    _c(9, cif or "")
    _c(10, ref or "")
    
    ws.row_dimensions[row].height = 18


def _desc_cuenta(cuenta: str) -> str:
    """Descripción textual simplificada de una cuenta contable española."""
    prefijos = {
        "400": "Proveedores",
        "401": "Proveedores, efectos comerciales a pagar",
        "410": "Acreedores por prestaciones de servicios",
        "472": "H.P. IVA Soportado",
        "600": "Compras de mercaderías",
        "601": "Compras de materias primas",
        "620": "Gastos en investigación y desarrollo",
        "621": "Arrendamientos y cánones",
        "622": "Reparaciones y conservación",
        "623": "Servicios de profesionales independientes",
        "624": "Transportes",
        "625": "Primas de seguros",
        "626": "Servicios bancarios y similares",
        "627": "Publicidad, propaganda y relaciones públicas",
        "628": "Suministros",
        "629": "Otros servicios",
        "631": "Otros tributos",
        "640": "Sueldos y salarios",
        "642": "Seguridad social a cargo de la empresa",
    }
    for prefix, desc in prefijos.items():
        if cuenta.startswith(prefix):
            return desc
    return f"Cuenta {cuenta}"


def _add_saldo_cuentas(wb, facturas: List[Dict]) -> None:
    """Hoja con saldo por cuenta contable (incluye subcuentas)."""
    ws = wb.create_sheet("Saldo por Cuenta")
    headers = ["Cuenta", "Subcuenta", "Descripción", "DEBE", "HABER", "Saldo"]
    for col, hdr in enumerate(headers, 1):
        _header_cell(ws, 1, col, hdr)

    cuentas: Dict[str, Dict] = {}
    for f in facturas:
        base  = float(f.get("base_imponible", 0) or 0)
        iva   = float(f.get("iva", 0) or 0)
        total = float(f.get("total", 0) or 0)
        es_rect = bool(f.get("es_rectificativa", 0))
        
        cg    = f.get("cuenta_gasto", "629000") or "629000"
        sg    = f.get("subcuenta_gasto", "") or ""
        cp    = f.get("cuenta_proveedor", "400000") or "400000"
        sp    = f.get("subcuenta_proveedor", "") or ""

        base_abs  = abs(base)
        iva_abs   = abs(iva)
        total_abs = abs(total)

        key     = f"{cg}|{sg}"
        key_iva = "472000|"
        key_prov = f"{cp}|{sp}"
        for k, cta, sub in [(key, cg, sg), (key_iva, "472000", ""), (key_prov, cp, sp)]:
            if k not in cuentas:
                cuentas[k] = {"cta": cta, "sub": sub, "debe": 0.0, "haber": 0.0}

        if not es_rect:
            # Factura normal: gasto/IVA al DEBE, proveedor al HABER
            cuentas[key]["debe"]     += base_abs
            cuentas[key_iva]["debe"] += iva_abs
            cuentas[key_prov]["haber"] += total_abs
        else:
            # Abono/rectificativa: gasto/IVA al HABER, proveedor al DEBE
            cuentas[key]["haber"]    += base_abs
            cuentas[key_iva]["haber"] += iva_abs
            cuentas[key_prov]["debe"] += total_abs

    for row_idx, (key, d) in enumerate(sorted(cuentas.items()), 2):
        saldo = round(d["debe"] - d["haber"], 2)
        ws.cell(row=row_idx, column=1, value=d["cta"])
        ws.cell(row=row_idx, column=2, value=d["sub"])
        ws.cell(row=row_idx, column=3, value=_desc_cuenta(d["cta"]))
        
        for col, val in [(4, d["debe"]), (5, d["haber"]), (6, saldo)]:
            c = ws.cell(row=row_idx, column=col, value=val)
            c.number_format = '#,##0.00 €'
            c.alignment = Alignment(horizontal="right")
            if col == 6 and saldo < 0:
                c.font = Font(color="C53030")

    for i, w in enumerate([12, 12, 35, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w