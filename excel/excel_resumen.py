# -*- coding: utf-8 -*-
"""
Excel Resumen entre fechas — V10.
Genera un Excel profesional con facturas filtradas por:
  - Rango de fechas
  - Proveedor
  - Categoría
  - Tipo de factura
Columnas: Fecha, Proveedor, CIF, Base, IVA, Total, Categoría, Tipo, Ruta, Nº Prov.
"""
from __future__ import annotations
import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ── Estilos ────────────────────────────────────────────────────────────────────

_HDR_FILL  = "1F4E79"
_HDR_FONT  = "FFFFFF"
_ALT_FILL  = "EBF8FF"
_TOTAL_FILL = "E8F5E9"


def _header_cell(ws, row: int, col: int, value: str) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=True, color=_HDR_FONT, size=10)
    cell.fill      = PatternFill("solid", fgColor=_HDR_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()


def _thin_border():
    s = Side(border_style="thin", color="CBD5E0")
    return Border(left=s, right=s, top=s, bottom=s)


def _money(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value or 0)
    cell.number_format = '#,##0.00 €'
    cell.alignment = Alignment(horizontal="right")
    return cell


# ── Exportador ────────────────────────────────────────────────────────────────

def _tipo_coste(cuenta_gasto: str) -> str:
    """F: Clasifica por primer dígito de cuenta: Gasto(6) / Compra(3-4) / Otro."""
    c = str(cuenta_gasto or "").strip()
    if c.startswith("6"): return "Gasto"
    if c.startswith(("3","4")): return "Compra"
    return "Otro"


def exportar_resumen(
    facturas: List[Dict],
    ruta_salida: Optional[str] = None,
    titulo: str = "Resumen de Facturas",
    incluir_tipo_coste: bool = True,
) -> str:
    """
    Genera un Excel con el resumen de facturas.
    Retorna la ruta del archivo generado.
    """
    if not OPENPYXL_OK:
        raise ImportError("openpyxl no instalado. Ejecuta: pip install openpyxl")

    if not ruta_salida:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_salida = f"Resumen_Facturas_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"

    # ── Título ─────────────────────────────────────────────────────────────────
    # Obtener nombre empresa para incluir en el informe
    _empresa_nom = ""
    try:
        from database.manager import DatabaseManager as _DM
        _empresa_nom = _DM().get_config_ui("empresa_nombre", "")
    except Exception:
        pass
    _titulo_comp = f"📋  {titulo}" + (f"  —  {_empresa_nom}" if _empresa_nom else "")

    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value     = _titulo_comp
    title_cell.font      = Font(bold=True, size=14, color=_HDR_FONT)
    title_cell.fill      = PatternFill("solid", fgColor="2B6CB0")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:K2")
    gen_cell = ws["A2"]
    gen_cell.value     = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    gen_cell.font      = Font(italic=True, size=9, color="4A5568")
    gen_cell.alignment = Alignment(horizontal="right")

    # ── Encabezados ────────────────────────────────────────────────────────────
    # B-FIX: Headers con cuentas/subcuentas, serie, nº factura, impresa
    headers = [
        "Fecha",            # 1
        "Proveedor",        # 2
        "Nº Prov.",         # 3
        "CIF",              # 4
        "Cta. Prov.",       # 5
        "Subcta. Prov.",    # 6
        "Base Imponible",   # 7
        "IVA",              # 8
        "Total",            # 9
        "Tipo Coste",       # 10
        "Cta. Gasto",       # 11
        "Subcta. Gasto",    # 12
        "Nº Factura",       # 13
        "Serie",            # 14
        "Rect.",            # 15
        "Impresa",          # 16
        "Tipo",             # 17
        "Ruta PDF",         # 18
    ]
    ws.merge_cells(f"A1:R1")
    title_cell = ws["A1"]
    title_cell.value     = _titulo_comp
    title_cell.font      = Font(bold=True, size=14, color=_HDR_FONT)
    title_cell.fill      = PatternFill("solid", fgColor="2B6CB0")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:R2")
    gen_cell = ws["A2"]
    gen_cell.value     = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    gen_cell.font      = Font(italic=True, size=9, color="4A5568")
    gen_cell.alignment = Alignment(horizontal="right")
    for col_idx, hdr in enumerate(headers, 1):
        _header_cell(ws, 3, col_idx, hdr)
    ws.row_dimensions[3].height = 24

    # ── Datos ─────────────────────────────────────────────────────────────────
    total_base = total_iva = total_total = 0.0

    for row_idx, f in enumerate(facturas, 4):
        alt = (row_idx % 2 == 0)
        fill = PatternFill("solid", fgColor=_ALT_FILL) if alt else None

        def _cell(col, val, num_fmt=None):
            c = ws.cell(row=row_idx, column=col, value=val)
            if fill:
                c.fill = fill
            c.border = _thin_border()
            if num_fmt:
                c.number_format = num_fmt
            return c

        _cell(1,  f.get("fecha", ""))
        _cell(2,  f.get("nombre_proveedor", ""))
        _cell(3,  f.get("numero_proveedor") or f.get("num_prov", ""))
        _cell(4,  f.get("cif_proveedor") or f.get("cif_prov", ""))
        _cell(5,  f.get("cuenta_proveedor", "400000"))
        _cell(6,  f.get("subcuenta_proveedor", ""))
        _money(ws, row_idx, 7,  f.get("base_imponible", 0))
        _money(ws, row_idx, 8,  f.get("iva", 0))
        _money(ws, row_idx, 9,  f.get("total", 0))
        _cell(10, _tipo_coste(f.get("cuenta_gasto", "")))
        _cell(11, f.get("cuenta_gasto", ""))
        _cell(12, f.get("subcuenta_gasto", ""))
        _cell(13, f.get("numero_factura", ""))
        _cell(14, f.get("serie_factura", "") or f.get("serie", ""))
        _cell(15, "SI" if f.get("es_rectificativa") else "NO")
        _cell(16, "SI" if f.get("impresa") else "NO")
        _cell(17, f.get("tipo_factura", ""))
        _cell(18, f.get("ruta_pdf") or f.get("ruta_archivo_final", ""))

        total_base  += float(f.get("base_imponible", 0) or 0)
        total_iva   += float(f.get("iva",            0) or 0)
        total_total += float(f.get("total",          0) or 0)

    # ── Fila de totales ────────────────────────────────────────────────────────
    tot_row = len(facturas) + 4
    ws.merge_cells(f"A{tot_row}:C{tot_row}")
    t_lbl = ws[f"A{tot_row}"]
    t_lbl.value = f"TOTAL  ({len(facturas)} facturas)"
    t_lbl.font  = Font(bold=True, size=10, color=_HDR_FONT)
    t_lbl.fill  = PatternFill("solid", fgColor=_HDR_FILL)
    t_lbl.alignment = Alignment(horizontal="center")

    for col, val in [(4, total_base), (5, total_iva), (6, total_total)]:
        c = ws.cell(row=tot_row, column=col, value=val)
        c.number_format = '#,##0.00 €'
        c.font = Font(bold=True, color=_HDR_FONT)
        c.fill = PatternFill("solid", fgColor=_HDR_FILL)
        c.alignment = Alignment(horizontal="right")

    # ── Anchos de columna ─────────────────────────────────────────────────────
    widths = [12, 30, 14, 14, 12, 14, 10, 22, 14, 14, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Hoja pivot por proveedor ───────────────────────────────────────────────
    _add_pivot_proveedor(wb, facturas)
    _add_pivot_categoria(wb, facturas)
    _add_pivot_serie(wb, facturas)
    _add_pivot_tipo(wb, facturas)
    _add_pivot_serie_tipo(wb, facturas)

    wb.save(ruta_salida)
    return ruta_salida


def _add_pivot_proveedor(wb, facturas: List[Dict]) -> None:
    """Añade hoja pivot: totales por proveedor."""
    ws = wb.create_sheet("Por Proveedor")
    headers = ["Proveedor", "CIF", "Nº Facturas", "Base Imponible", "IVA", "Total"]
    for col, hdr in enumerate(headers, 1):
        _header_cell(ws, 1, col, hdr)

    # Agrupar
    grupos: Dict[str, Dict] = {}
    for f in facturas:
        k = f.get("nombre_proveedor", "DESCONOCIDO") or "DESCONOCIDO"
        if k not in grupos:
            grupos[k] = {
                "cif": f.get("cif_proveedor") or f.get("cif_prov", ""),
                "n": 0, "base": 0.0, "iva": 0.0, "total": 0.0
            }
        grupos[k]["n"]     += 1
        grupos[k]["base"]  += float(f.get("base_imponible", 0) or 0)
        grupos[k]["iva"]   += float(f.get("iva", 0) or 0)
        grupos[k]["total"] += float(f.get("total", 0) or 0)

    for row_idx, (nombre, d) in enumerate(
            sorted(grupos.items(), key=lambda x: -x[1]["total"]), 2):
        ws.cell(row=row_idx, column=1, value=nombre)
        ws.cell(row=row_idx, column=2, value=d["cif"])
        ws.cell(row=row_idx, column=3, value=d["n"])
        for col, val in [(4, d["base"]), (5, d["iva"]), (6, d["total"])]:
            c = ws.cell(row=row_idx, column=col, value=val)
            c.number_format = '#,##0.00 €'

    for i, w in enumerate([30, 14, 12, 16, 12, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Gráfico de barras
    if len(grupos) > 0 and len(grupos) <= 30:
        try:
            chart = BarChart()
            chart.title = "Gasto Total por Proveedor"
            chart.style = 10
            chart.y_axis.title = "€"
            data = Reference(ws, min_col=6, min_row=1, max_row=len(grupos)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(grupos)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.shape = 4
            ws.add_chart(chart, f"H2")
        except Exception:
            pass


def _add_pivot_categoria(wb, facturas: List[Dict]) -> None:
    """Añade hoja pivot: totales por categoría."""
    ws = wb.create_sheet("Por Categoría")
    headers = ["Categoría", "Nº Facturas", "Total"]
    for col, hdr in enumerate(headers, 1):
        _header_cell(ws, 1, col, hdr)

    grupos: Dict[str, Dict] = {}
    for f in facturas:
        k = f.get("categoria", "SIN CATEGORÍA") or "SIN CATEGORÍA"
        if k not in grupos:
            grupos[k] = {"n": 0, "total": 0.0}
        grupos[k]["n"]     += 1
        grupos[k]["total"] += float(f.get("total", 0) or 0)

    for row_idx, (cat, d) in enumerate(
            sorted(grupos.items(), key=lambda x: -x[1]["total"]), 2):
        ws.cell(row=row_idx, column=1, value=cat)
        ws.cell(row=row_idx, column=2, value=d["n"])
        c = ws.cell(row=row_idx, column=3, value=d["total"])
        c.number_format = '#,##0.00 €'

    for i, w in enumerate([30, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Gráfico circular
    if len(grupos) > 0 and len(grupos) <= 20:
        try:
            chart = PieChart()
            chart.title = "Distribución por Categoría"
            data = Reference(ws, min_col=3, min_row=1, max_row=len(grupos)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(grupos)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "E2")
        except Exception:
            pass


def _add_pivot_serie(wb, facturas: List[Dict]) -> None:
    """Hoja pivot: totales por serie de factura."""
    ws = wb.create_sheet("Por Serie")
    headers = ["Serie", "Nº Facturas", "Base Imponible", "IVA", "Total"]
    for col, hdr in enumerate(headers, 1):
        _header_cell(ws, 1, col, hdr)

    grupos: Dict[str, Dict] = {}
    for f in facturas:
        k = f.get("serie_factura") or f.get("serie", "") or "(Sin serie)"
        if k not in grupos:
            grupos[k] = {"n": 0, "base": 0.0, "iva": 0.0, "total": 0.0}
        grupos[k]["n"]     += 1
        grupos[k]["base"]  += float(f.get("base_imponible", 0) or 0)
        grupos[k]["iva"]   += float(f.get("iva", 0) or 0)
        grupos[k]["total"] += float(f.get("total", 0) or 0)

    for row_idx, (nombre, d) in enumerate(
            sorted(grupos.items(), key=lambda x: -x[1]["base"]), 2):
        ws.cell(row=row_idx, column=1, value=nombre)
        ws.cell(row=row_idx, column=2, value=d["n"])
        for col, val in [(3, d["base"]), (4, d["iva"]), (5, d["total"])]:
            c = ws.cell(row=row_idx, column=col, value=val)
            c.number_format = '#,##0.00 €'

    # Fila de total
    tot_row = len(grupos) + 2
    ws.cell(row=tot_row, column=1, value="TOTAL")
    ws.cell(row=tot_row, column=1).font = Font(bold=True)
    ws.cell(row=tot_row, column=2, value=sum(d["n"] for d in grupos.values()))
    for col, key in [(3, "base"), (4, "iva"), (5, "total")]:
        c = ws.cell(row=tot_row, column=col,
                    value=sum(d[key] for d in grupos.values()))
        c.number_format = '#,##0.00 €'
        c.font = Font(bold=True)

    for i, w in enumerate([22, 12, 16, 12, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Gráfico de barras por serie
    if 0 < len(grupos) <= 30:
        try:
            chart = BarChart()
            chart.title = "Base Imponible por Serie"
            chart.style = 10
            chart.y_axis.title = "€"
            data = Reference(ws, min_col=3, min_row=1, max_row=len(grupos) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(grupos) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "G2")
        except Exception:
            pass


def _add_pivot_tipo(wb, facturas: List[Dict]) -> None:
    """Hoja pivot: totales por tipo de factura."""
    ws = wb.create_sheet("Por Tipo")
    headers = ["Tipo Factura", "Nº Facturas", "Base Imponible", "IVA", "Total"]
    for col, hdr in enumerate(headers, 1):
        _header_cell(ws, 1, col, hdr)

    grupos: Dict[str, Dict] = {}
    for f in facturas:
        k = f.get("tipo_factura") or "(Sin tipo)"
        if k not in grupos:
            grupos[k] = {"n": 0, "base": 0.0, "iva": 0.0, "total": 0.0}
        grupos[k]["n"]     += 1
        grupos[k]["base"]  += float(f.get("base_imponible", 0) or 0)
        grupos[k]["iva"]   += float(f.get("iva", 0) or 0)
        grupos[k]["total"] += float(f.get("total", 0) or 0)

    for row_idx, (nombre, d) in enumerate(
            sorted(grupos.items(), key=lambda x: -x[1]["base"]), 2):
        ws.cell(row=row_idx, column=1, value=nombre)
        ws.cell(row=row_idx, column=2, value=d["n"])
        for col, val in [(3, d["base"]), (4, d["iva"]), (5, d["total"])]:
            c = ws.cell(row=row_idx, column=col, value=val)
            c.number_format = '#,##0.00 €'

    # Fila total
    tot_row = len(grupos) + 2
    ws.cell(row=tot_row, column=1, value="TOTAL")
    ws.cell(row=tot_row, column=1).font = Font(bold=True)
    ws.cell(row=tot_row, column=2, value=sum(d["n"] for d in grupos.values()))
    for col, key in [(3, "base"), (4, "iva"), (5, "total")]:
        c = ws.cell(row=tot_row, column=col,
                    value=sum(d[key] for d in grupos.values()))
        c.number_format = '#,##0.00 €'
        c.font = Font(bold=True)

    for i, w in enumerate([22, 12, 16, 12, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    if 0 < len(grupos) <= 20:
        try:
            chart = PieChart()
            chart.title = "Distribución por Tipo de Factura"
            data = Reference(ws, min_col=3, min_row=1, max_row=len(grupos) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(grupos) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "G2")
        except Exception:
            pass


def _add_pivot_serie_tipo(wb, facturas: List[Dict]) -> None:
    """Hoja pivot: combinación Serie × Tipo de factura."""
    ws = wb.create_sheet("Serie × Tipo")
    headers = ["Serie", "Tipo Factura", "Nº Facturas", "Base Imponible", "IVA", "Total"]
    for col, hdr in enumerate(headers, 1):
        _header_cell(ws, 1, col, hdr)

    grupos: Dict[tuple, Dict] = {}
    for f in facturas:
        serie = f.get("serie_factura") or f.get("serie", "") or "(Sin serie)"
        tipo  = f.get("tipo_factura") or "(Sin tipo)"
        k = (serie, tipo)
        if k not in grupos:
            grupos[k] = {"n": 0, "base": 0.0, "iva": 0.0, "total": 0.0}
        grupos[k]["n"]     += 1
        grupos[k]["base"]  += float(f.get("base_imponible", 0) or 0)
        grupos[k]["iva"]   += float(f.get("iva", 0) or 0)
        grupos[k]["total"] += float(f.get("total", 0) or 0)

    for row_idx, ((serie, tipo), d) in enumerate(
            sorted(grupos.items(), key=lambda x: (x[0][0], -x[1]["base"])), 2):
        fill = PatternFill("solid", fgColor=_ALT_FILL) if row_idx % 2 == 0 else None
        for col_i, val in enumerate([serie, tipo, d["n"]], 1):
            c = ws.cell(row=row_idx, column=col_i, value=val)
            if fill:
                c.fill = fill
            c.border = _thin_border()
        for col_i, val in [(4, d["base"]), (5, d["iva"]), (6, d["total"])]:
            c = ws.cell(row=row_idx, column=col_i, value=val)
            c.number_format = '#,##0.00 €'
            if fill:
                c.fill = fill
            c.border = _thin_border()

    for i, w in enumerate([22, 20, 12, 16, 12, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def exportar_informe_avanzado(
    dimensiones: List[str],
    fecha_desde: str = None,
    fecha_hasta: str = None,
    filtros: Optional[Dict] = None,
    ruta_salida: Optional[str] = None,
) -> str:
    """
    Genera un Excel con un informe avanzado usando cualquier combinación de
    columnas de la base de datos (serie_factura, tipo_factura, categoria,
    nombre_proveedor, cuenta_gasto, nombre_regla, etc.).

    Ejemplo de uso:
        exportar_informe_avanzado(
            dimensiones=["serie_factura", "tipo_factura"],
            fecha_desde="2026-01-01",
            fecha_hasta="2026-12-31",
        )
    """
    if not OPENPYXL_OK:
        raise ImportError("openpyxl no instalado. pip install openpyxl")

    from database.manager import DatabaseManager as _DM
    db = _DM()
    filas = db.obtener_informe_avanzado(dimensiones, fecha_desde, fecha_hasta, filtros)

    if not ruta_salida:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dims_str = "_".join(dimensiones[:2])
        ruta_salida = f"Informe_Avanzado_{dims_str}_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe Avanzado"

    titulo = f"Informe Avanzado — {' × '.join(dimensiones)}"
    n_cols = len(dimensiones) + 4  # dims + n_facturas + base + iva + total
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    tc = ws["A1"]
    tc.value     = titulo
    tc.font      = Font(bold=True, size=13, color=_HDR_FONT)
    tc.fill      = PatternFill("solid", fgColor="2B6CB0")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{get_column_letter(n_cols)}2")
    sc = ws["A2"]
    sc.value     = f"Período: {fecha_desde or '—'} → {fecha_hasta or '—'}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sc.font      = Font(italic=True, size=9, color="4A5568")
    sc.alignment = Alignment(horizontal="right")

    # Cabeceras
    _LABELS = {
        "serie_factura": "Serie", "tipo_factura": "Tipo", "categoria": "Categoría",
        "nombre_proveedor": "Proveedor", "cuenta_gasto": "Cta. Gasto",
        "subcuenta_gasto": "Subcta. Gasto", "cuenta_proveedor": "Cta. Prov.",
        "numero_proveedor": "Nº Prov.", "cif_proveedor": "CIF",
        "nombre_regla": "Regla", "id_regla_aplicada": "ID Regla",
        "es_rectificativa": "Rectif.", "procesada_desde_correo": "Desde Correo",
    }
    headers = [_LABELS.get(d, d) for d in dimensiones] + [
        "Nº Facturas", "Base Imponible", "IVA", "Total"]
    for col, hdr in enumerate(headers, 1):
        _header_cell(ws, 3, col, hdr)
    ws.row_dimensions[3].height = 22

    total_base = total_iva = total_total = 0.0
    for row_idx, fila in enumerate(filas, 4):
        alt = (row_idx % 2 == 0)
        fill = PatternFill("solid", fgColor=_ALT_FILL) if alt else None
        for col_i, dim in enumerate(dimensiones, 1):
            c = ws.cell(row=row_idx, column=col_i, value=fila.get(dim, ""))
            if fill:
                c.fill = fill
            c.border = _thin_border()
        base_col = len(dimensiones) + 1
        ws.cell(row=row_idx, column=base_col, value=fila.get("num_facturas", 0))
        for offset, key in [(1, "total_base"), (2, "total_iva"), (3, "total_total")]:
            val = float(fila.get(key, 0) or 0)
            c = ws.cell(row=row_idx, column=base_col + offset, value=val)
            c.number_format = '#,##0.00 €'
            if fill:
                c.fill = fill
            c.border = _thin_border()
        total_base  += float(fila.get("total_base", 0) or 0)
        total_iva   += float(fila.get("total_iva", 0) or 0)
        total_total += float(fila.get("total_total", 0) or 0)

    # Fila de totales
    tot_row = len(filas) + 4
    ws.merge_cells(f"A{tot_row}:{get_column_letter(len(dimensiones))}{tot_row}")
    tl = ws[f"A{tot_row}"]
    tl.value = f"TOTAL ({len(filas)} agrupaciones)"
    tl.font  = Font(bold=True, color=_HDR_FONT)
    tl.fill  = PatternFill("solid", fgColor=_HDR_FILL)
    tl.alignment = Alignment(horizontal="center")
    base_col = len(dimensiones) + 2
    for offset, val in [(0, total_base), (1, total_iva), (2, total_total)]:
        c = ws.cell(row=tot_row, column=base_col + offset, value=val)
        c.number_format = '#,##0.00 €'
        c.font = Font(bold=True, color=_HDR_FONT)
        c.fill = PatternFill("solid", fgColor=_HDR_FILL)

    # Ancho columnas
    for i in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22 if i <= len(dimensiones) else 16

    wb.save(ruta_salida)
    return ruta_salida
