# -*- coding: utf-8 -*-
"""
importador_proveedores.py — importador batch sin UI.
Reemplaza la versión que usaba tkinter.messagebox con logging estándar.
"""
import openpyxl
import logging

log = logging.getLogger(__name__)


class ImportadorProveedores:
    def __init__(self, db_manager):
        self.db = db_manager

    def importar_desde_excel(self, ruta_excel: str) -> dict:
        """
        Importa proveedores desde un archivo Excel.
        Devuelve {'importados': int, 'errores': list[str]}.
        """
        importados = 0
        errores = []
        try:
            wb = openpyxl.load_workbook(ruta_excel, data_only=True)
            ws = wb.active
            cabeceras = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                fila = {cab: str(val or "").strip()
                        for cab, val in zip(cabeceras, row) if cab}
                nombre = fila.get("nombre") or fila.get("nombre comercial", "")
                if not nombre:
                    continue
                datos = {
                    "nombre":           nombre,
                    "numero_proveedor": fila.get("numero_proveedor") or fila.get("nº proveedor", ""),
                    "cuenta_gasto":     fila.get("cuenta_gasto") or fila.get("cuenta", ""),
                    "categoria":        fila.get("categoria", "VARIOS"),
                    "cif_nif":          fila.get("cif_nif") or fila.get("cif", ""),
                    "email":            fila.get("email", ""),
                    "iban":             fila.get("iban", ""),
                }
                try:
                    self.db.insertar_proveedor(datos)
                    importados += 1
                except Exception as exc:
                    errores.append(f"{nombre}: {exc}")
        except Exception as exc:
            errores.append(f"Error general: {exc}")
            log.error("Error importando proveedores: %s", exc)
        return {"importados": importados, "errores": errores}
