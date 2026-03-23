# -*- coding: utf-8 -*-
"""
importar_excel.py — PyQt5 nativo.
Importa proveedores desde un archivo Excel con vista previa y mapeo de columnas.
Reemplaza la versión Tkinter.

Exporta:
    ImportarProveedoresDialog(db, parent=None)   — QDialog
    abrir_importar_proveedores(parent, callback_reload=None)  — función de compatibilidad
"""
from __future__ import annotations
import sys, os
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QGroupBox,
    QLabel, QLineEdit, QComboBox, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QFileDialog,
    QMessageBox, QProgressBar, QCheckBox, QApplication
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal

BTN_P = ("QPushButton{background:#1F4E79;color:white;border-radius:4px;"
         "padding:6px 14px;font-weight:bold;}QPushButton:hover{background:#2B6CB0;}")
BTN_S = ("QPushButton{background:#EDF2F7;color:#2D3748;border:1px solid #CBD5E0;"
         "border-radius:4px;padding:6px 14px;}QPushButton:hover{background:#E2E8F0;}")
TBL   = ("QTableWidget{border:1px solid #E2E8F0;border-radius:4px;}"
         "QHeaderView::section{background:#F7FAFC;border-bottom:1px solid #CBD5E0;"
         "padding:5px 8px;font-weight:bold;}")
COMBO = "QComboBox{border:1px solid #CBD5E0;border-radius:4px;padding:4px 6px;background:white;}"


def _get_categorias():
    try:
        from core.config_loader import get_config
        return get_config().categories or []
    except Exception:
        pass
    try:
        from config import CATEGORIAS
        return CATEGORIAS
    except Exception:
        return ["COMERCIAL Y POSTVENTA EXTERNAS", "COMUNES EXTERNOS", "GESTORIA", "SEAT", "VARIOS"]


CAMPOS = [
    ("nombre",           "* Nombre del Proveedor", True),
    ("numero_proveedor", "* Nº de Proveedor",       True),
    ("cuenta_gasto",     "Cuenta de Gasto",         False),
    ("cif_nif",          "CIF / NIF",               False),
    ("categoria",        "Categoría / Carpeta",     False),
    ("email",            "Email",                   False),
    ("iban",             "IBAN",                    False),
]


class ImportarProveedoresDialog(QDialog):
    """
    Diálogo de 3 pasos para importar proveedores desde Excel:
      1. Seleccionar archivo y hoja
      2. Mapear columnas del Excel a los campos del proveedor
      3. Vista previa y confirmación
    """
    def __init__(self, db=None, parent=None):
        super().__init__(parent)
        from database import DatabaseManager
        self.db = db or DatabaseManager()
        self.setWindowTitle("📥  Importar Proveedores desde Excel")
        self.setMinimumSize(860, 680)
        self.setWindowFlag(Qt.WindowContextHelpButtonHint, False)
        self._workbook  = None
        self._hoja      = None
        self._cols      = []   # nombres de columnas del Excel
        self._mapeos    = {}   # campo → índice de columna
        self._combos    = {}   # campo → QComboBox de mapeo
        self._filas     = []   # lista de dicts para importar
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 16, 16, 14)
        lay.setSpacing(12)

        titulo = QLabel("📥  IMPORTAR PROVEEDORES DESDE EXCEL")
        titulo.setStyleSheet("font-size:14px;font-weight:bold;color:#1F4E79;")
        lay.addWidget(titulo)

        # ── Paso 1: selección de archivo y hoja ──────────────────────────────
        box1 = QGroupBox("1. Seleccionar Archivo y Hoja")
        box1.setStyleSheet("QGroupBox{font-weight:bold;border:1px solid #CBD5E0;border-radius:6px;margin-top:8px;padding-top:12px;}")
        b1 = QHBoxLayout(box1)

        self.lbl_archivo = QLabel("No se ha seleccionado ningún archivo")
        self.lbl_archivo.setStyleSheet("color:#718096;")
        btn_sel = QPushButton("📁  Seleccionar Excel")
        btn_sel.setStyleSheet(BTN_S)
        btn_sel.setFixedHeight(32)
        btn_sel.clicked.connect(self._seleccionar_archivo)
        self.cmb_hoja = QComboBox()
        self.cmb_hoja.setStyleSheet(COMBO)
        self.cmb_hoja.setEnabled(False)
        self.cmb_hoja.setFixedWidth(160)
        self.cmb_hoja.currentIndexChanged.connect(self._cargar_preview)
        b1.addWidget(btn_sel)
        b1.addWidget(self.lbl_archivo, 1)
        b1.addWidget(QLabel("Hoja:"))
        b1.addWidget(self.cmb_hoja)
        lay.addWidget(box1)

        # ── Paso 2: mapeo de columnas ──────────────────────────────────────────
        box2 = QGroupBox("2. Mapear Columnas del Excel")
        box2.setStyleSheet("QGroupBox{font-weight:bold;border:1px solid #CBD5E0;border-radius:6px;margin-top:8px;padding-top:12px;}")
        b2 = QVBoxLayout(box2)
        b2.addWidget(QLabel("Indica qué columna del Excel corresponde a cada campo "
                            "(* = obligatorio):"))
        self._form_mapeo = QFormLayout()
        self._form_mapeo.setLabelAlignment(Qt.AlignRight)
        self._form_mapeo.setSpacing(6)
        for campo, etiqueta, requerido in CAMPOS:
            cmb = QComboBox()
            cmb.setStyleSheet(COMBO)
            cmb.setFixedHeight(28)
            cmb.addItem("— No mapear —", None)
            self._combos[campo] = cmb
            self._form_mapeo.addRow(f"{etiqueta}:", cmb)
        b2.addLayout(self._form_mapeo)
        lay.addWidget(box2)

        # ── Paso 3: vista previa ─────────────────────────────────────────────
        box3 = QGroupBox("3. Vista Previa y Confirmación")
        box3.setStyleSheet("QGroupBox{font-weight:bold;border:1px solid #CBD5E0;border-radius:6px;margin-top:8px;padding-top:12px;}")
        b3 = QVBoxLayout(box3)

        preview_bar = QHBoxLayout()
        btn_prev = QPushButton("👁  Cargar Vista Previa")
        btn_prev.setStyleSheet(BTN_S)
        btn_prev.setFixedHeight(32)
        btn_prev.clicked.connect(self._generar_preview)
        self.lbl_preview_info = QLabel("")
        self.lbl_preview_info.setStyleSheet("color:#4A5568;")
        preview_bar.addWidget(btn_prev)
        preview_bar.addWidget(self.lbl_preview_info, 1)
        b3.addLayout(preview_bar)

        self.tbl_preview = QTableWidget()
        self.tbl_preview.setStyleSheet(TBL)
        self.tbl_preview.setAlternatingRowColors(True)
        self.tbl_preview.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl_preview.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_preview.verticalHeader().setVisible(False)
        b3.addWidget(self.tbl_preview)
        lay.addWidget(box3)

        # ── Barra final ────────────────────────────────────────────────────────
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        lay.addWidget(self.progress)

        bar = QHBoxLayout()
        bar.addStretch()
        btn_cancel = QPushButton("Cancelar"); btn_cancel.setStyleSheet(BTN_S)
        btn_cancel.clicked.connect(self.reject)
        self.btn_importar = QPushButton("✅  IMPORTAR PROVEEDORES")
        self.btn_importar.setStyleSheet(BTN_P)
        self.btn_importar.setEnabled(False)
        self.btn_importar.clicked.connect(self._importar)
        bar.addWidget(btn_cancel); bar.addWidget(self.btn_importar)
        lay.addLayout(bar)

    def _seleccionar_archivo(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo Excel", "",
            "Excel (*.xlsx *.xls);;Todos los archivos (*)")
        if not path:
            return
        try:
            import openpyxl
            self._workbook = openpyxl.load_workbook(path, data_only=True)
            self.lbl_archivo.setText(os.path.basename(path))
            self.lbl_archivo.setStyleSheet("color:#276749;font-weight:bold;")
            self.cmb_hoja.clear()
            for nombre in self._workbook.sheetnames:
                self.cmb_hoja.addItem(nombre)
            self.cmb_hoja.setEnabled(True)
            self._cargar_preview()
        except ImportError:
            QMessageBox.critical(self, "openpyxl no instalado",
                                 "Instala openpyxl:\n  pip install openpyxl")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"No se pudo abrir el archivo:\n{exc}")

    def _cargar_preview(self):
        if not self._workbook:
            return
        nombre_hoja = self.cmb_hoja.currentText()
        if not nombre_hoja:
            return
        try:
            ws = self._workbook[nombre_hoja]
            self._hoja = ws
            # Primera fila = cabeceras
            primera_fila = next(ws.iter_rows(max_row=1, values_only=True), [])
            self._cols = [str(c or "").strip() for c in primera_fila]
            # Actualizar combos de mapeo
            for campo, cmb in self._combos.items():
                cmb.clear()
                cmb.addItem("— No mapear —", None)
                for i, col in enumerate(self._cols):
                    if col:
                        cmb.addItem(col, i)
                # Auto-mapear por nombre similar
                for i, col in enumerate(self._cols):
                    if campo.lower().replace("_", "") in col.lower().replace(" ", "").replace("_", ""):
                        cmb.setCurrentIndex(i + 1)  # +1 por el "No mapear"
                        break
            self.lbl_preview_info.setText(
                f"Hoja cargada: {len(self._cols)} columnas. "
                f"Haz clic en 'Cargar Vista Previa' para revisar.")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Error al cargar la hoja:\n{exc}")

    def _generar_preview(self):
        if not self._hoja:
            QMessageBox.information(self, "Sin datos", "Selecciona un archivo y hoja primero.")
            return
        # Recopilar mapeos
        self._mapeos = {}
        for campo, cmb in self._combos.items():
            idx = cmb.currentData()
            if idx is not None:
                self._mapeos[campo] = idx

        campos_requeridos = [c for c, _, req in CAMPOS if req]
        for cr in campos_requeridos:
            if cr not in self._mapeos:
                QMessageBox.warning(self, "Mapeo incompleto",
                                    f"El campo obligatorio «{cr}» no está mapeado.")
                return

        # Leer filas de datos
        cols_preview = [f for f, _, _ in CAMPOS if f in self._mapeos]
        self.tbl_preview.setColumnCount(len(cols_preview))
        self.tbl_preview.setHorizontalHeaderLabels(
            [next(e for c, e, _ in CAMPOS if c == f) for f in cols_preview])
        self.tbl_preview.setRowCount(0)
        self.tbl_preview.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)

        self._filas = []
        for row in self._hoja.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            fila = {}
            for campo, col_idx in self._mapeos.items():
                val = row[col_idx] if col_idx < len(row) else ""
                fila[campo] = str(val or "").strip()
            # Solo incluir si nombre no está vacío
            if not fila.get("nombre"):
                continue
            self._filas.append(fila)
            trow = self.tbl_preview.rowCount()
            self.tbl_preview.insertRow(trow)
            for ci, campo in enumerate(cols_preview):
                self.tbl_preview.setItem(trow, ci, QTableWidgetItem(fila.get(campo, "")))

        n = len(self._filas)
        self.lbl_preview_info.setText(f"{n} proveedor{'es' if n != 1 else ''} listos para importar.")
        self.btn_importar.setEnabled(n > 0)

    def _importar(self):
        if not self._filas:
            return
        self.progress.setVisible(True)
        self.progress.setMaximum(len(self._filas))
        self.btn_importar.setEnabled(False)
        importados = 0
        errores = []
        cat_default = _get_categorias()[0] if _get_categorias() else "VARIOS"
        for i, fila in enumerate(self._filas):
            self.progress.setValue(i + 1)
            QApplication.processEvents()
            fila.setdefault("cuenta_gasto",     "")
            fila.setdefault("categoria",        cat_default)
            fila.setdefault("numero_proveedor", f"AUTO-{i+1:04d}")
            try:
                self.db.insertar_proveedor(fila)
                importados += 1
            except Exception as exc:
                errores.append(f"{fila.get('nombre', '?')}: {exc}")
        self.progress.setVisible(False)
        msg = f"✅ Importados: {importados}\n"
        if errores:
            msg += f"⚠️ Errores ({len(errores)}):\n" + "\n".join(errores[:10])
        QMessageBox.information(self, "Importación finalizada", msg)
        if importados > 0:
            self.accept()
        else:
            self.btn_importar.setEnabled(True)


# ─────────────────────────────────────────────────────────────────────────────
# API de compatibilidad
# ─────────────────────────────────────────────────────────────────────────────

def abrir_importar_proveedores(parent_widget=None, callback_reload=None):
    """Compatibilidad con el código antiguo que llamaba a esta función."""
    from database import DatabaseManager
    db = DatabaseManager()
    dlg = ImportarProveedoresDialog(db=db, parent=parent_widget)
    if dlg.exec_() == QDialog.Accepted and callback_reload:
        callback_reload()
