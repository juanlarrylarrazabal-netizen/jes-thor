# -*- coding: utf-8 -*-
"""
Exportador Excel del módulo laboral.
Genera informes de:
  - Costes laborales (nóminas por empleado y mes)
  - Asistencia (fichajes mensuales)
  - Plantilla personalizable: carga un Excel existente y rellena sus campos

Reutiliza el patrón de excel/excel_resumen.py (openpyxl, mismos estilos).
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

from core.logging_config import get_logger

log = get_logger("laboral.informes")

_MESES_ES = {
    1:"Enero", 2:"Febrero", 3:"Marzo", 4:"Abril",
    5:"Mayo", 6:"Junio", 7:"Julio", 8:"Agosto",
    9:"Septiembre", 10:"Octubre", 11:"Noviembre", 12:"Diciembre"
}

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# Estilos consistentes con el resto de la app
_HDR_FILL  = "1F4E79"
_ALT_FILL  = "EFF6FF"
_HDR_FONT  = {"bold": True, "color": "FFFFFF", "size": 10}
_BODY_FONT = {"size": 9}


def _hdr_style(ws, row, cols):
    fill = PatternFill("solid", fgColor=_HDR_FILL)
    font = Font(bold=True, color="FFFFFF", size=10)
    aln  = Alignment(horizontal="center", vertical="center")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = aln


def _alt_fill(ws, row, cols):
    fill = PatternFill("solid", fgColor=_ALT_FILL)
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).fill = fill


# ── Informe de costes laborales ────────────────────────────────────────────────

def exportar_costes_laborales(
    db_laboral,
    anio: int,
    mes_inicio: int = 1,
    mes_fin: int = 12,
    carpeta_salida: str = "./informes",
) -> Optional[str]:
    """
    Genera Excel con costes laborales del período indicado.
    Columnas: Empleado | Mes | Salario Base | Complementos |
              SS Empleado | IRPF | Líquido | SS Empresa | Coste Total
    """
    if not OPENPYXL_OK:
        log.error("openpyxl no disponible")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Costes Laborales"

    # Título
    ws.merge_cells("A1:I1")
    ws["A1"].value = f"INFORME DE COSTES LABORALES — {anio}"
    ws["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill  = PatternFill("solid", fgColor=_HDR_FILL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Generado
    ws["A2"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font  = Font(italic=True, size=8, color="666666")
    ws.row_dimensions[2].height = 16

    # Cabecera
    headers = ["Empleado", "NIF", "Mes", "Salario Base", "Complementos",
               "SS Empleado", "IRPF", "Líquido", "SS Empresa", "Coste Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col).value = h
    _hdr_style(ws, 3, len(headers))
    ws.row_dimensions[3].height = 20

    # Datos
    nominas = db_laboral.obtener_nominas(anio=anio)
    row_num = 4
    totales = {k: 0.0 for k in ["salario_base","complementos","ss_empleado",
                                  "irpf","liquido","ss_empresa","coste_empresa"]}

    for n in nominas:
        if not (mes_inicio <= (n.get("mes") or 0) <= mes_fin):
            continue
        mes_str = _MESES_ES.get(n.get("mes", 0), str(n.get("mes", "")))
        vals = [
            n.get("nombre_empleado", ""),
            n.get("nif_empleado", ""),
            mes_str,
            n.get("salario_base", 0) or 0,
            n.get("complementos", 0) or 0,
            n.get("ss_empleado", 0) or 0,
            n.get("irpf", 0) or 0,
            n.get("liquido", 0) or 0,
            n.get("ss_empresa", 0) or 0,
            n.get("coste_empresa", 0) or 0,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            if col >= 4:
                cell.number_format = '#,##0.00 €'
        if row_num % 2 == 0:
            _alt_fill(ws, row_num, len(headers))

        for k, campo in enumerate(["salario_base","complementos","ss_empleado",
                                     "irpf","liquido","ss_empresa","coste_empresa"], 3):
            totales[campo] += float(n.get(campo, 0) or 0)
        row_num += 1

    # Fila totales
    ws.cell(row=row_num, column=1).value = "TOTALES"
    ws.cell(row=row_num, column=1).font  = Font(bold=True)
    for col, campo in enumerate(["salario_base","complementos","ss_empleado",
                                   "irpf","liquido","ss_empresa","coste_empresa"], 4):
        cell = ws.cell(row=row_num, column=col, value=totales[campo])
        cell.number_format = '#,##0.00 €'
        cell.font = Font(bold=True)

    # Anchos de columna
    anchos = [28, 14, 12, 14, 14, 14, 12, 14, 14, 14]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    # Guardar
    Path(carpeta_salida).mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = str(Path(carpeta_salida) / f"Costes_Laborales_{anio}_{ts}.xlsx")
    wb.save(ruta)
    log.info("Informe costes laborales guardado: %s", ruta)
    return ruta


# ── Informe de asistencia ──────────────────────────────────────────────────────

def exportar_asistencia_mensual(
    db_laboral,
    anio: int,
    mes: int,
    carpeta_salida: str = "./informes",
) -> Optional[str]:
    """Genera Excel de asistencia para todos los empleados en un mes."""
    if not OPENPYXL_OK:
        log.error("openpyxl no disponible")
        return None

    from laboral.fichajes.zkteco import generar_informe_asistencia

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Asistencia {_MESES_ES.get(mes,mes)} {anio}"

    ws.merge_cells("A1:H1")
    ws["A1"].value = f"INFORME DE ASISTENCIA — {_MESES_ES.get(mes,mes).upper()} {anio}"
    ws["A1"].font  = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill  = PatternFill("solid", fgColor=_HDR_FILL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Empleado", "Días Trabajados", "Horas Totales",
               "Retrasos", "Jornadas Incompletas", "Días Laborables"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col).value = h
    _hdr_style(ws, 2, len(headers))

    empleados = db_laboral.obtener_empleados(solo_activos=True)
    row_num = 3
    for emp in empleados:
        inf = generar_informe_asistencia(db_laboral, emp["id"], anio, mes)
        vals = [
            inf["empleado"],
            inf["dias_trabajados"],
            inf["horas_trabajadas"],
            inf["retrasos"],
            inf["jornadas_incompletas"],
            inf["dias_laborables"],
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row_num, column=col, value=v)
        if row_num % 2 == 0:
            _alt_fill(ws, row_num, len(headers))
        row_num += 1

    anchos = [28, 16, 14, 10, 20, 16]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    Path(carpeta_salida).mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = str(Path(carpeta_salida) / f"Asistencia_{anio}_{mes:02d}_{ts}.xlsx")
    wb.save(ruta)
    log.info("Informe asistencia guardado: %s", ruta)
    return ruta


# ── Rellenar plantilla Excel existente ────────────────────────────────────────

def rellenar_plantilla_excel(
    ruta_plantilla: str,
    mapa_campos: Dict[str, str],
    datos: Dict[str, object],
    ruta_salida: str = None,
) -> Optional[str]:
    """
    Rellena una plantilla Excel existente de la empresa.
    
    mapa_campos: {"A5": "salario_base", "B5": "complementos", ...}
                 Las claves son celdas Excel, los valores son claves en `datos`.
    datos:       {"salario_base": 1500.00, "complementos": 200.00, ...}
    """
    if not OPENPYXL_OK:
        log.error("openpyxl no disponible")
        return None

    try:
        wb = openpyxl.load_workbook(ruta_plantilla)
        ws = wb.active

        for celda, campo in mapa_campos.items():
            if campo in datos and datos[campo] is not None:
                ws[celda] = datos[campo]
                log.debug("Plantilla: celda %s → %s = %s",
                          celda, campo, datos[campo])

        if ruta_salida is None:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            base = Path(ruta_plantilla)
            ruta_salida = str(base.parent / f"{base.stem}_rellenado_{ts}{base.suffix}")

        wb.save(ruta_salida)
        log.info("Plantilla Excel rellenada: %s", ruta_salida)
        return ruta_salida
    except Exception as e:
        log.error("Error rellenando plantilla Excel: %s", e)
        return None
