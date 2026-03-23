# -*- coding: utf-8 -*-
"""
Integración con sistemas de fichaje ZKTeco.
Soporta:
  A) Importación desde CSV/Excel exportados por el software ZKTeco
  B) Conexión directa al terminal por red (zkpy / pyzk)

Reutiliza: core.logging_config, storage.filesystem.safe_name
"""
from __future__ import annotations
import csv
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import List, Dict, Optional, Tuple

from core.logging_config import get_logger

log = get_logger("laboral.fichajes")


# ── Importación desde CSV/Excel ────────────────────────────────────────────────

class ImportadorFichajesCSV:
    """
    Importa fichajes desde los archivos CSV/Excel que genera el software ZKTeco.

    Formato CSV típico ZKTeco:
        ID_Empleado, Nombre, Departamento, Fecha_Hora, Tipo_Evento
        1, "GARCIA LOPEZ JUAN", "TALLER", "2026-03-01 08:02:35", 0
    """

    # Columnas esperadas (varias variantes de nombre)
    _COL_ID_EMP  = ["id", "id_empleado", "emp id", "userid", "user id"]
    _COL_NOMBRE  = ["nombre", "name", "emp name", "nombre completo"]
    _COL_FECHA   = ["fecha_hora", "datetime", "check time", "fecha", "time"]
    _COL_TIPO    = ["tipo", "type", "verify type", "estado", "event"]

    def __init__(self, db_laboral=None):
        if db_laboral is None:
            from laboral.db_laboral import LaboralDB
            db_laboral = LaboralDB()
        self.db = db_laboral

    def _normalizar_cols(self, headers: List[str]) -> Dict[str, int]:
        """Mapea nombre de columna → índice, de forma flexible."""
        mapping = {}
        headers_low = [h.lower().strip() for h in headers]
        for campo, variantes in [
            ("id_emp", self._COL_ID_EMP),
            ("nombre", self._COL_NOMBRE),
            ("fecha",  self._COL_FECHA),
            ("tipo",   self._COL_TIPO),
        ]:
            for variante in variantes:
                if variante in headers_low:
                    mapping[campo] = headers_low.index(variante)
                    break
        return mapping

    def _parsear_fecha_hora(self, texto: str) -> Tuple[str, str]:
        """Devuelve (fecha_iso, hora_hhmm) desde varios formatos."""
        texto = texto.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S",
                    "%d/%m/%Y %H:%M", "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M"):
            try:
                dt = datetime.strptime(texto, fmt)
                return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M")
            except ValueError:
                continue
        log.debug("No se pudo parsear fecha: %r", texto)
        return date.today().isoformat(), "00:00"

    def importar_csv(self, ruta_csv: str,
                     separador: str = ",",
                     encoding: str = "utf-8") -> Dict:
        """Importa fichajes desde CSV. Devuelve estadísticas."""
        importados = duplicados = errores = 0
        empleados  = self.db.obtener_empleados()

        try:
            with open(ruta_csv, encoding=encoding, errors="replace") as fh:
                reader = csv.reader(fh, delimiter=separador)
                headers = next(reader)
                col_map = self._normalizar_cols(headers)
                log.info("CSV fichajes: columnas detectadas %s", col_map)

                # Agrupar por empleado+fecha para calcular entrada/salida
                registros: Dict[Tuple, List[str]] = {}
                for row in reader:
                    if not any(c.strip() for c in row):
                        continue
                    try:
                        nombre_col = col_map.get("nombre")
                        fecha_col  = col_map.get("fecha")
                        if fecha_col is None:
                            continue

                        nombre_raw = row[nombre_col].strip() if nombre_col else ""
                        fecha_raw  = row[fecha_col].strip()
                        fecha_iso, hora = self._parsear_fecha_hora(fecha_raw)

                        # Buscar empleado por nombre
                        emp = (self.db.buscar_empleado_por_nombre(nombre_raw)
                               if nombre_raw else None)
                        emp_id = emp["id"] if emp else None

                        clave = (emp_id, fecha_iso)
                        registros.setdefault(clave, []).append(hora)

                    except Exception as e:
                        errores += 1
                        log.debug("Error en fila CSV: %s", e)

            # Insertar fichajes agrupados (entrada = mínimo, salida = máximo)
            for (emp_id, fecha_iso), horas in registros.items():
                horas_ord = sorted(horas)
                entrada = horas_ord[0]
                salida  = horas_ord[-1] if len(horas_ord) > 1 else None

                minutos = 0
                if salida and salida != entrada:
                    try:
                        h1 = datetime.strptime(entrada, "%H:%M")
                        h2 = datetime.strptime(salida, "%H:%M")
                        minutos = int((h2 - h1).total_seconds() / 60)
                    except Exception:
                        pass

                fila = {
                    "empleado_id":         emp_id,
                    "fecha":               fecha_iso,
                    "hora_entrada":        entrada,
                    "hora_salida":         salida,
                    "minutos_trabajados":  minutos,
                    "origen":              "zkteco_csv",
                }
                result = self.db.insertar_fichaje(fila)
                if result:
                    importados += 1
                else:
                    duplicados += 1

        except FileNotFoundError:
            log.error("Fichero CSV no encontrado: %s", ruta_csv)
            return {"error": "Fichero no encontrado"}
        except Exception as e:
            log.error("Error importando CSV fichajes: %s", e)
            return {"error": str(e)}

        log.info("Fichajes CSV importados: %d nuevos, %d duplicados, %d errores",
                 importados, duplicados, errores)
        return {"importados": importados, "duplicados": duplicados,
                "errores": errores}

    def importar_excel(self, ruta_excel: str) -> Dict:
        """Importa fichajes desde Excel ZKTeco."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(ruta_excel, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return {"error": "Excel vacío"}

            headers = [str(c or "").lower().strip() for c in rows[0]]
            col_map = self._normalizar_cols(headers)

            import tempfile
            import csv as _csv
            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv',
                                             delete=False, encoding='utf-8') as tmp:
                writer = _csv.writer(tmp)
                writer.writerow(rows[0])
                for row in rows[1:]:
                    writer.writerow([str(c or "") for c in row])
                tmp_path = tmp.name

            return self.importar_csv(tmp_path)
        except ImportError:
            return {"error": "openpyxl no instalado"}
        except Exception as e:
            log.error("Error importando Excel fichajes: %s", e)
            return {"error": str(e)}


# ── Conexión directa ZKTeco ────────────────────────────────────────────────────

class ConectorZKTeco:
    """
    Conecta directamente al terminal ZKTeco por red (puerto 4370).
    Requiere: pip install pyzk
    Devuelve gracefully si pyzk no está instalado.
    """

    def __init__(self, ip: str, puerto: int = 4370,
                 password: int = 0, timeout: int = 10):
        self.ip      = ip
        self.puerto  = puerto
        self.password = password
        self.timeout = timeout
        self._zk = None

    def conectar(self) -> bool:
        try:
            from zk import ZK
            zk = ZK(self.ip, port=self.puerto, timeout=self.timeout,
                    password=self.password, force_udp=False, ommit_ping=False)
            self._zk = zk.connect()
            log.info("ZKTeco conectado: %s:%d", self.ip, self.puerto)
            return True
        except ImportError:
            log.warning("pyzk no instalado. Instalar con: pip install pyzk")
            return False
        except Exception as e:
            log.error("Error conectando a ZKTeco %s:%d → %s",
                      self.ip, self.puerto, e)
            return False

    def desconectar(self) -> None:
        try:
            if self._zk:
                self._zk.disconnect()
                self._zk = None
        except Exception:
            pass

    def descargar_fichajes(self, db_laboral=None) -> Dict:
        """Descarga todos los registros de asistencia del terminal."""
        if not self._zk:
            return {"error": "No conectado"}
        importados = errores = 0
        try:
            if db_laboral is None:
                from laboral.db_laboral import LaboralDB
                db_laboral = LaboralDB()

            empleados_zk = self._zk.get_users()
            empleados_bd = db_laboral.obtener_empleados()
            # Mapeo NIF del ZKTeco → empleado BD por nombre
            mapa_id: Dict[str, int] = {}
            for uz in empleados_zk:
                emp = db_laboral.buscar_empleado_por_nombre(uz.name or "")
                if emp:
                    mapa_id[str(uz.uid)] = emp["id"]

            registros = self._zk.get_attendance()
            agrupados: Dict[Tuple, List] = {}
            for r in registros:
                emp_id = mapa_id.get(str(r.user_id))
                fecha_iso = r.timestamp.strftime("%Y-%m-%d")
                hora_str  = r.timestamp.strftime("%H:%M")
                agrupados.setdefault((emp_id, fecha_iso), []).append(hora_str)

            for (emp_id, fecha_iso), horas in agrupados.items():
                horas_ord = sorted(horas)
                entrada = horas_ord[0]
                salida  = horas_ord[-1] if len(horas_ord) > 1 else None
                minutos = 0
                if salida and salida != entrada:
                    h1 = datetime.strptime(entrada, "%H:%M")
                    h2 = datetime.strptime(salida, "%H:%M")
                    minutos = int((h2 - h1).total_seconds() / 60)

                result = db_laboral.insertar_fichaje({
                    "empleado_id":        emp_id,
                    "fecha":              fecha_iso,
                    "hora_entrada":       entrada,
                    "hora_salida":        salida,
                    "minutos_trabajados": minutos,
                    "origen":             "zkteco_api",
                    "dispositivo":        f"{self.ip}:{self.puerto}",
                })
                importados += 1 if result else 0

            log.info("ZKTeco: %d fichajes descargados", importados)
            return {"importados": importados, "errores": errores}
        except Exception as e:
            log.error("Error descargando fichajes ZKTeco: %s", e)
            return {"error": str(e), "importados": importados}
        finally:
            self.desconectar()

    def borrar_registros_terminal(self) -> bool:
        """Borra los registros del terminal tras importarlos (opcional)."""
        try:
            if self._zk:
                self._zk.clear_attendance()
                log.info("Registros borrados del terminal ZKTeco")
                return True
        except Exception as e:
            log.error("Error borrando registros ZKTeco: %s", e)
        return False


# ── Generador de informes de fichajes ─────────────────────────────────────────

def generar_informe_asistencia(db_laboral, empleado_id: int,
                                anio: int, mes: int) -> Dict:
    """
    Genera un informe de asistencia mensual para un empleado.
    Devuelve dict con métricas listas para mostrar o exportar a Excel.
    """
    from calendar import monthrange
    _, dias_mes = monthrange(anio, mes)

    fecha_desde = f"{anio}-{mes:02d}-01"
    fecha_hasta = f"{anio}-{mes:02d}-{dias_mes}"

    fichajes = db_laboral.obtener_fichajes(
        empleado_id=empleado_id,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
    )
    resumen = db_laboral.resumen_horas_mes(empleado_id, anio, mes)
    empleado = db_laboral.obtener_empleado(empleado_id)

    total_minutos = resumen.get("total_minutos") or 0
    horas_trabajadas = total_minutos / 60

    return {
        "empleado":          f"{empleado.get('nombre','')} {empleado.get('apellidos','')}",
        "periodo":           f"{mes:02d}/{anio}",
        "dias_trabajados":   resumen.get("dias_trabajados", 0),
        "horas_trabajadas":  round(horas_trabajadas, 2),
        "retrasos":          resumen.get("retrasos", 0),
        "jornadas_incompletas": resumen.get("jornadas_incompletas", 0),
        "dias_laborables":   dias_mes,
        "detalle":           fichajes,
    }


# ── Descarga de empleados desde terminal ZKTeco ───────────────────────────────

class DescargaEmpleadosZKTeco:
    """
    Descarga la lista de empleados registrados en el terminal ZKTeco
    y la sincroniza con la base de datos del módulo laboral.
    Requiere: pip install pyzk
    """

    def __init__(self, ip: str, puerto: int = 4370, timeout: int = 10):
        self.ip      = ip
        self.puerto  = puerto
        self.timeout = timeout

    def obtener_empleados_terminal(self) -> Dict:
        """
        Conecta al terminal y devuelve lista de empleados.
        Cada entrada: {"uid": int, "nombre": str, "user_id": str}
        """
        try:
            from zk import ZK
        except ImportError:
            log.warning("pyzk no instalado. Instalar con: pip install pyzk")
            return {"error": "pyzk no instalado", "empleados": []}

        try:
            zk   = ZK(self.ip, port=self.puerto, timeout=self.timeout,
                      force_udp=False, ommit_ping=False)
            conn = zk.connect()
            usuarios = conn.get_users()
            conn.disconnect()

            empleados = [
                {"uid": u.uid, "nombre": u.name or "", "user_id": str(u.user_id or "")}
                for u in usuarios
            ]
            log.info("ZKTeco: %d empleados descargados de %s:%d",
                     len(empleados), self.ip, self.puerto)
            return {"empleados": empleados, "total": len(empleados)}
        except Exception as e:
            log.error("Error descargando empleados de ZKTeco: %s", e)
            return {"error": str(e), "empleados": []}

    def importar_a_bd(self, db_laboral, empleados_terminal: list,
                      solo_nuevos: bool = True) -> Dict:
        """
        Importa empleados del terminal a la BD laboral.
        solo_nuevos=True: solo añade los que no existen (por nombre).
        Devuelve estadísticas de la importación.
        """
        importados = ya_existian = 0
        empleados_bd = {
            f"{e['nombre']} {e['apellidos']}".upper(): e
            for e in db_laboral.obtener_empleados()
        }

        for emp_zk in empleados_terminal:
            nombre_zk = (emp_zk.get("nombre") or "").strip().upper()
            if not nombre_zk:
                continue
            if nombre_zk in empleados_bd:
                ya_existian += 1
                continue
            if solo_nuevos:
                # Intentar separar nombre y apellidos (convención ZKTeco: APELLIDOS NOMBRE)
                partes = nombre_zk.split()
                if len(partes) >= 2:
                    apellidos = " ".join(partes[:-1]).title()
                    nombre    = partes[-1].title()
                else:
                    apellidos = nombre_zk.title()
                    nombre    = "—"
                try:
                    db_laboral.insertar_empleado({
                        "nombre":    nombre,
                        "apellidos": apellidos,
                        "estado":    "activo",
                        "notas":     f"Importado desde ZKTeco (uid={emp_zk.get('uid')})",
                    })
                    importados += 1
                    log.info("Empleado ZKTeco importado: %s %s", nombre, apellidos)
                except Exception as e:
                    log.warning("Error importando empleado ZKTeco %s: %s", nombre_zk, e)

        log.info("Importación ZKTeco: %d nuevos, %d ya existían",
                 importados, ya_existian)
        return {"importados": importados, "ya_existian": ya_existian}


# ── Importador ZKTime ─────────────────────────────────────────────────────────

class ImportadorZKTime:
    """
    Importa datos exportados desde el software ZKTime (software oficial ZKTeco).

    ZKTime puede exportar:
    - Empleados: CSV/Excel con columnas ID, Nombre, Departamento, etc.
    - Fichajes:  CSV/Excel con columnas ID, Nombre, Fecha/Hora, Tipo
    - Informes:  varios formatos

    Esta clase detecta automáticamente el tipo de exportación e importa
    lo que encuentre al módulo laboral.
    """

    # Columnas típicas ZKTime para empleados
    _EMP_COLS = {
        "id":          ["emp no", "emp id", "id", "employee no", "numero"],
        "nombre":      ["name", "nombre", "emp name", "employee name", "full name"],
        "departamento":["department", "dept", "departamento"],
        "cargo":       ["position", "cargo", "puesto", "job title"],
    }

    # Columnas típicas ZKTime para fichajes
    _FICH_COLS = {
        "id_emp":  ["emp no", "emp id", "id", "employee no"],
        "nombre":  ["name", "nombre", "emp name"],
        "fecha":   ["check time", "datetime", "fecha_hora", "time", "fecha"],
        "tipo":    ["verify type", "check type", "tipo"],
    }

    def __init__(self, db_laboral=None):
        if db_laboral is None:
            from laboral.db_laboral import LaboralDB
            db_laboral = LaboralDB()
        self.db = db_laboral

    def _detectar_tipo(self, headers: List[str]) -> str:
        """Detecta si el archivo es de empleados o de fichajes según las cabeceras."""
        heads_low = [h.lower().strip() for h in headers]
        # Si tiene columna de hora/check time → fichajes
        for fich_var in ["check time", "time", "datetime", "fecha_hora"]:
            if any(fich_var in h for h in heads_low):
                return "fichajes"
        # Si tiene columna de nombre sin hora → empleados
        for emp_var in ["name", "nombre", "emp name"]:
            if any(emp_var in h for h in heads_low):
                return "empleados"
        return "desconocido"

    def _col_idx(self, headers_low: List[str], variantes: List[str]) -> Optional[int]:
        for v in variantes:
            for i, h in enumerate(headers_low):
                if v in h:
                    return i
        return None

    def importar_csv(self, ruta: str, encoding: str = "utf-8",
                     separador: str = ",") -> Dict:
        """Importa CSV exportado de ZKTime. Auto-detecta empleados o fichajes."""
        resultado = {"tipo": "desconocido", "importados": 0, "errores": 0}
        try:
            import csv
            with open(ruta, encoding=encoding, errors="replace") as fh:
                reader = csv.reader(fh, delimiter=separador)
                headers = next(reader)
                filas   = list(reader)
        except FileNotFoundError:
            return {"error": "Archivo no encontrado"}
        except Exception as e:
            return {"error": str(e)}

        tipo = self._detectar_tipo(headers)
        resultado["tipo"] = tipo

        if tipo == "empleados":
            r = self._importar_empleados_filas(headers, filas)
        elif tipo == "fichajes":
            from laboral.fichajes.zkteco import ImportadorFichajesCSV
            imp = ImportadorFichajesCSV(self.db)
            r   = imp.importar_csv(ruta, separador=separador, encoding=encoding)
        else:
            return {"error": f"No se reconoce el formato (cabeceras: {headers[:5]})"}

        resultado.update(r)
        return resultado

    def importar_excel(self, ruta: str) -> Dict:
        """Importa Excel exportado de ZKTime. Auto-detecta empleados o fichajes."""
        try:
            import openpyxl
            wb   = openpyxl.load_workbook(ruta, read_only=True)
            ws   = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not rows:
                return {"error": "Excel vacío"}
        except ImportError:
            return {"error": "openpyxl no instalado"}
        except Exception as e:
            return {"error": str(e)}

        headers = [str(c or "").strip() for c in rows[0]]
        filas   = [[str(c or "").strip() for c in row] for row in rows[1:]]
        tipo    = self._detectar_tipo(headers)

        if tipo == "empleados":
            return self._importar_empleados_filas(headers, filas)
        elif tipo == "fichajes":
            # Convertir a CSV temporal y reutilizar ImportadorFichajesCSV
            import tempfile, csv
            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv',
                                             delete=False, encoding='utf-8') as tmp:
                writer = csv.writer(tmp)
                writer.writerow(headers)
                for f in filas:
                    writer.writerow(f)
                tmp_path = tmp.name
            from laboral.fichajes.zkteco import ImportadorFichajesCSV
            return ImportadorFichajesCSV(self.db).importar_csv(tmp_path)
        else:
            return {"error": f"Formato no reconocido (cabeceras: {headers[:5]})"}

    def _importar_empleados_filas(self, headers: List[str],
                                   filas: List[List[str]]) -> Dict:
        """Importa filas de empleados ZKTime a la BD laboral."""
        heads_low = [h.lower().strip() for h in headers]
        col_nombre = self._col_idx(heads_low, ["name", "nombre", "emp name", "employee name"])
        col_dept   = self._col_idx(heads_low, ["department", "dept", "departamento"])
        col_cargo  = self._col_idx(heads_low, ["position", "cargo", "puesto"])

        if col_nombre is None:
            return {"error": "No se encontró columna de nombre en los datos"}

        # Empleados ya en BD (para no duplicar)
        existentes = {
            f"{e['nombre']} {e['apellidos']}".upper()
            for e in self.db.obtener_empleados()
        }

        importados = ya_existian = errores = 0
        for fila in filas:
            if not any(c.strip() for c in fila):
                continue
            try:
                nombre_raw = fila[col_nombre].strip() if col_nombre < len(fila) else ""
                if not nombre_raw:
                    continue

                # ZKTime suele poner "APELLIDO NOMBRE" → intentar separar
                partes = nombre_raw.upper().split()
                if len(partes) >= 2:
                    apellidos = " ".join(partes[:-1]).title()
                    nombre    = partes[-1].title()
                else:
                    apellidos = nombre_raw.title()
                    nombre    = "—"

                clave = f"{apellidos} {nombre}".upper().strip()
                if any(clave in ex for ex in existentes):
                    ya_existian += 1
                    continue

                datos = {
                    "nombre":    nombre,
                    "apellidos": apellidos,
                    "estado":    "activo",
                }
                if col_dept and col_dept < len(fila):
                    datos["notas"] = f"Dpto: {fila[col_dept]}"
                if col_cargo and col_cargo < len(fila):
                    datos["categoria"] = fila[col_cargo]

                self.db.insertar_empleado(datos)
                existentes.add(clave)
                importados += 1
            except Exception as e:
                errores += 1
                log.debug("Error importando empleado ZKTime: %s", e)

        log.info("ZKTime empleados: %d importados, %d ya existían, %d errores",
                 importados, ya_existian, errores)
        return {"importados": importados, "ya_existian": ya_existian, "errores": errores}
